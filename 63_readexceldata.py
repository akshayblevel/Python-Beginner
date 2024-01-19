import openpyxl

wb = openpyxl.load_workbook('test.xlsx')

ws = wb.active

print('Total Rows: '+str(ws.max_row)+'. Total Columns: '+str(ws.max_column))

my_list = list()
for value in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
    my_list.append(value)

for i in my_list:
    print(i)
